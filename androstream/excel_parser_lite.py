"""
Lightweight Excel Parser for Android
Does not use pandas - uses openpyxl directly (lighter)
"""

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelParserLite:
    """Lightweight Excel parser without pandas dependency"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.file_path = None
        self.rows_data = []
    
    def load_excel(self, file_path):
        """Load Excel file"""
        if not HAS_OPENPYXL:
            return False, "openpyxl not available. Excel import disabled for this build."
        
        try:
            self.file_path = file_path
            self.workbook = load_workbook(file_path)
            self.worksheet = self.workbook.active
            
            # Parse all rows
            self.rows_data = []
            headers = [cell.value for cell in self.worksheet[1]]
            
            for row in self.worksheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                self.rows_data.append(row_dict)
            
            return True, f"Loaded {len(self.rows_data)} rows from Excel"
        except Exception as e:
            return False, f"Error loading Excel: {str(e)}"
    
    def get_all_rows(self):
        """Get all parsed rows"""
        return self.rows_data
    
    def get_preview(self, limit=10):
        """Get preview of first N rows"""
        return self.rows_data[:limit]


# For backward compatibility
ExcelParser = ExcelParserLite
